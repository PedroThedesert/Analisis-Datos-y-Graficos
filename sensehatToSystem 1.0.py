""" Gestión de datos climáticos, distribución de datos.  """

"""
Nombre: sensehatToSystem 1.0.py
Descripcion: Información, gráficas de datos, guardado en base de datos, distribución de plantillas
Autor: Pedro Gutiérrez Herrero
Fecha Creacion: 08/04/2019
"""

# -*- coding: utf-8 -*-

import sqlite3
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from tkinter import filedialog
from openpyxl import Workbook


#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
#+++++++++++++++++++      CLASES             +++++++++++++++++++++++++++++++++
#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++


class DDBB:
    """ Provee métodos y atributos a la clase de Base de Datos  """       
    # Clase para realizar interactuaciones con la base de datos
        
    def __init__(self, name=None):
        """ Inicializamos la clase de Base de Datos  """    
        # Constructor de clase        
        self.conn = None
        self.cursor = None

        if name:
            self.open(name)

    
    def abrirDB(self,name):
        """ Establece conexión con la Base de datos  """
        #  Abre la conexión con la base de datos 
        # name es el nombre de la base de datos a abrir
        try:
            self.conn = sqlite3.connect(name);
            self.cursor = self.conn.cursor()
            print("Conexión Realizada a la base de datos\n")
        except sqlite3.Error:
            print("Error de conexión a la base de datos\n")

    
    def cerrarDB(self):
        """ Cierra la conexión con la base de datos  """
        # Cierra la base de datos
    
        if self.conn:
            self.conn.commit()
            self.cursor.close()
            self.conn.close()
            print ("Conexión con la base de datos cerrada con Éxito\n")

    
    def conseguirDatos(self,table,columns,limit=None):
        """ Obtiene los datos al realizar una consulta  """    
        # Función para obtener datos de una base de datos 
        #  table es el nombre de la base de datos usada
        #  columns es el string de columnas para capturar
        #  limit establece un límite de datos a capturar
    
    
        query = "SELECT {0} from {1};".format(columns,table)       
        self.cursor.execute(query)

        # capturar datos
        rows = self.cursor.fetchall()

        return rows[len(rows)-limit if limit else 0:]


    def borraDatos_Tabla(self,table):
        """ Borra todos los datos contenidos en la tabla  """     
        ## Borra los datos de la tabla
        
        query = "Delete from {0};".format(table) 
        self.cursor.execute(query)
        print ("BORRADO DE DATOS REALIZADO CORRECTAMENTE\n")


    def realizaConsulta(self,sql):
        """ Realiza una consulta a la tabla  """
        # Función para realizar consultas sql
        # En caso de no querer escribir o obtener datos
        # sql es una declaración valida SQL en formato string
        
        self.cursor.execute(sql)
        return(self.cursor.fetchall())


class Graphics:
    """ Provee métodos y atributos a la clase para realizar Gráficos """           
    # Clase para realizar Graficas
        
    def __init__(self, listaSeleccionada, tipo):
        """ Inicializamos atributos para la clase Gráficos """  
        # Constructor de clase        
        self.listaSeleccionada = listaSeleccionada
        print("datos de la lista", listaSeleccionada)
        self.tipo = tipo
        
        plt.clf()
        plt.cla()
        plt.close()
        
        plt.plot(listaSeleccionada)
        
    def confAxis(self):
        """ Configuro los ejes y titulos para cada tipo de gráfico """  
        # Configuro los ejes y titulos para cada tipo de gráfico
    
        if self.tipo == "Temperatura":            
            plt.axis([0,23,-10,30])
            plt.ylabel('Temperatura')
            plt.xlabel('Horas')
        elif self.tipo == 'Humedad':
            plt.axis([0,23, 50,100])
            plt.ylabel('HUMEDAD')
            plt.xlabel('Horas')
        elif self.tipo == 'Presion':
            plt.axis([0,23,700,1100])
            plt.ylabel('PRESIÓN')
            plt.xlabel('Horas')
    
        plt.show()


#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
#++++++++++++++++++++    FUNCIONES    ++++++++++++++++++++++++++++++++++++++++
#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++


def obtenerListas_Separadas(lista):
    """ Mete los datos del excel en listas separadas para analísis posterior """  
    # Me coloca los datos extraídos de sql en cuatro listas separadas por dato
    
    # Inicializo listas
    
    listaTemperatura.clear()
    listaHumedad.clear()
    listaPresion.clear()
    listaFecha.clear()

    for x in lista:
        listaFecha.append(x[2])
        listaTemperatura.append(x[3])
        listaHumedad.append(x[4])
        listaPresion.append(x[5])
        
def excelToVariable():
    """ Mete el contenido de excel en una sola variable """  
    # Devuelve una variable con el contenido de todas las filas del archivo excel seleccionado    
    FilaCompuesta = [] 
    
    
    # Simulado
    #┘FILE_PATH = 'E:/Programacion/Python/Curso Cifo/Materiales/Materiales Curso/Serverpython/Modul-1/Sesion 35 - Desarrollo Proyecto/proyecto/pruebas/cargadorexcel.xlsx'
    
 

    FILE_PATH = filedialog.askopenfilename()
    print(FILE_PATH)
    
    
    excel_document = load_workbook(FILE_PATH)    
    excel_document.sheetnames 
    sheet = excel_document["DATOS"] 
    

    
    for row in sheet.iter_rows(min_row=2): # for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
     
         # unir todas las filas de excel
         Fila = [(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value)]
         FilaCompuesta = FilaCompuesta + Fila  
        
    return (FilaCompuesta)
 
def is_empty(data_structure):
    """ Devuelve un booleano dependiendo de si hay o no hay algo en la estructura a analizar """    
    # Comprueba si tenemos datos en una lista    
    if data_structure:
        return False
    else:
        print("La tabla se encuentra vacía. Inserte datos para generar gráficos.\n")
        return True


#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
#+++++++++++++++    SECUENCIA PRINCIPAL    +++++++++++++++++++++++++++++++++++
#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

# INICIALIZACIÓN
    
namedb = 'climadatos.db'
tabledb = 'datosCiudades'
columnsdb = ("Usuario, Ciudad, Fecha, Temperatura, Humedad, Presion, FechaExacta" )

# Listas personalizadas


listaTemperatura =  []
listaHumedad =  []
listaPresion =  []
listaFecha = []

# Instancias

midb = DDBB()



# Menú Aplicación

condicion = True

print("""
++++++++++++++++++++++++++++++++++++++++++++++++++++++++
+++++++++++      SENSEHAT TO SYSTEM             ++++++++
++++++++++++++++++++++++++++++++++++++++++++++++++++++++
""")




while (condicion == True):
    print("""¿Que quieres hacer? Escribe una opción

 1) AÑADIR A BASE DE DATOS
 2) GENERACIÓN DE GRÁFICO
 3) GENERACIÓN EXCEL
 4) BORRADO DE BASE DE DATOS   
 5) VISUALIZAR DATOS DE LA TABLA  
 6) SALIR APP""")
    opcion = input("")
    if opcion == '1':
        
        # abre la base de datos
        midb.abrirDB(namedb)

        print("SELECCIONA FICHERO EXCEL A CARGAR EN LA BASE DE DATOS\n") 
        
        datasetFromExcel = excelToVariable()
        #print ("----DATASET EXCEL LISTA A TUPLA------") 
        datasetFromExcel = tuple(datasetFromExcel)
        
        #print(type(datasetFromExcel))
        midb.cursor.executemany("INSERT INTO datosCiudades VALUES (?, ?, ?, ? , ?, ?, ?)", datasetFromExcel)
        
        print ("DATOS GUARDADOS CORRECTAMENTE EN BASE DE DATOS\n")
        
        # cierra la base de datos
        midb.cerrarDB()
    elif opcion == '2':
        print("GENERACIÓN DE GRÁFICO")
        # abre la base de datos
        midb.abrirDB(namedb)

        filtroCiudad = str(input("Introduce una Ciudad para filtrar en la consulta: Ejemplo 'Esplugues'\n"))
        filtroDia = str(input("Introduce un Día para filtrar en la consulta: Ejemplo '21/05/2019'\n"))
        
        # Prueba asignación directa
        #filtroCiudad = str('Esplugues')
        #filtroDia = str('20/05/2019')
        
        
        
        sqlQuery = "SELECT * FROM datosCiudades WHERE Ciudad= '{}' and FechaExacta= '{}'".format (filtroCiudad, filtroDia)
        
        dataConsulta = midb.realizaConsulta(sqlQuery)
        
        
        if is_empty(dataConsulta) == True: 
            break    
        
        # Si activamos veremos la consulta y los datos
        #printDataset()
        
        lista = []
        
        lista = dataConsulta
        
            
        obtenerListas_Separadas(lista)    
        
        # si lo activamos veremos las listas de ta, humedad y presión por separado
        #printlistaSeparadas()
        
        # GRÁFICAS VARIADAS 
                
        Grap_temp = Graphics(listaTemperatura, 'Temperatura')
        Grap_temp.confAxis()
        Grap_hum = Graphics(listaHumedad, 'Humedad')
        Grap_hum.confAxis()
        Grap_pres = Graphics(listaPresion, 'Presion')
        Grap_pres.confAxis()

        # cierra la base de datos
        midb.cerrarDB()
        

    elif opcion == '3':    
        print("GENERACIÓN EXCEL\n")
        # abre la base de datos
        midb.abrirDB(namedb)

        # Filtros Seleccionables
        filtroCiudad = str(input("Introduce una Ciudad para obtener sus datos: Ejemplo 'Esplugues'\n"))
        filtroDia = str(input("Introduce un Día para obtener sus datos: Ejemplo '21/05/2019'\n"))
        
        # Prueba asignación directa
        #filtroCiudad = str('Esplugues')
        #filtroDia = str('20/05/2019')
                       
        sqlQuery = "SELECT * FROM datosCiudades WHERE Ciudad= '{}' and FechaExacta= '{}'".format (filtroCiudad, filtroDia)
        
        dataConsulta = midb.realizaConsulta(sqlQuery)

        
        # cierra la base de datos
        midb.cerrarDB()

        # Genera Excel
        wb = Workbook()
        
        # selecciona nombre para el fichero
        #dest_filename = str(input("Elige nombre de fichero para guardar reporte. Ejemplo: '20190501 - Barcelona.xlsx'\n")) 
        dest_filename = filtroCiudad + filtroDia.replace("/", "_") + ".xlsx"
        print("El archivo ha sido guardado en el directorio E:\Programacion\Python\Curso Cifo\Materiales\Materiales Curso\Serverpython\Modul-1\Sesion 35 - Desarrollo Proyecto\proyecto\pruebas con el nombre ",  dest_filename)
        
        ws1 = wb.active       
        ws1.title = "Reporte"

        i = 1
        filaexcel = 0
        
        # Rellena los datos de la lista a la pestaña Reporte
        for i, dataToCell in enumerate(dataConsulta):
            columnExcel = 0
            filaexcel = filaexcel+1
            for item in dataToCell:              
                columnExcel = columnExcel + 1
                ws1.cell(row=filaexcel, column = columnExcel).value = item
        
      
        wb.save(filename = dest_filename)
        print ("\nFICHERO EXCEL GENERADO CORRECTAMENTE\n")
        
    elif opcion == '4':    
        print("BORRADO DE BASE DE DATOS\n")
        # abre la base de datos
        midb.abrirDB(namedb)

        #Borrar los datos de la tabla
        midb.borraDatos_Tabla(tabledb)
        
                
        # cierra la base de datos
        midb.cerrarDB()

    elif opcion == '5':    
        print("VISUALIZAR DATOS DE LA TABLA\n")
        # abre la base de datos
        midb.abrirDB(namedb)

        dataRecuperada = midb.conseguirDatos(tabledb,columnsdb)

        print ("")
        print ("DATOS RECUPERADOS DE LA TABLA\n")
        print ("+++++++++++++++-------------------+++++++++++++++")
        print (dataRecuperada)
        print ("+++++++++++++++-------------------+++++++++++++++")
        print ("")

        # cierra la base de datos
        midb.cerrarDB()
        
    elif opcion == '6':    
        condicion = False    
        























